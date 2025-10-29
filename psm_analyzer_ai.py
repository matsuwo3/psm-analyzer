"""
PSM分析自動レポートツール（AI統合版）
Claude API を活用してビジネスインサイトを生成する
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import anthropic
import os
from datetime import datetime
from dotenv import load_dotenv

# .envファイルから環境変数を読み込む
load_dotenv()


# ページ設定
st.set_page_config(
    page_title="PSM分析自動レポートツール（AI版）",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# セッション状態の初期化
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'analysis_complete' not in st.session_state:
    st.session_state.analysis_complete = False
if 'api_key' not in st.session_state:
    st.session_state.api_key = os.environ.get('ANTHROPIC_API_KEY', '')

# パスワード認証
APP_PASSWORD = "matsuo1234"

def check_password():
    """パスワード認証を行う"""
    if st.session_state.authenticated:
        return True

    st.title("🔐 PSM分析ツール（AI版）")
    st.markdown("---")
    st.markdown("### ログイン")
    st.info("このアプリケーションを使用するには、パスワードを入力してください。")

    password = st.text_input("パスワード", type="password", key="password_input")

    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("ログイン", type="primary", use_container_width=True):
            if password == APP_PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("❌ パスワードが正しくありません")

    return False


# ========== ユーティリティ関数 ==========

def safe_int(value):
    """安全に整数に変換"""
    try:
        return int(value)
    except:
        return value


def calculate_cumulative_distribution(data: np.ndarray, ascending: bool = True) -> tuple:
    """
    累積分布を計算する

    Args:
        data: 価格データの配列
        ascending: True=昇順（「高い」「高すぎる」）、False=降順（「安い」「お買い得」）

    Returns:
        tuple: (価格の配列, 累積パーセンテージの配列)
    """
    sorted_data = np.sort(data)
    n = len(sorted_data)

    if ascending:
        # 昇順累積分布（価格が上がるほど増加）
        cumulative = np.arange(1, n + 1) / n * 100
    else:
        # 降順累積分布（価格が上がるほど減少）
        cumulative = (n - np.arange(0, n)) / n * 100

    return sorted_data, cumulative


def find_intersection(x1: np.ndarray, y1: np.ndarray,
                     x2: np.ndarray, y2: np.ndarray) -> tuple:
    """
    2つの曲線の交点を線形補間で求める

    Args:
        x1, y1: 曲線1の座標
        x2, y2: 曲線2の座標

    Returns:
        tuple: (交点のx座標, 交点のy座標) または (None, None)
    """
    # 共通の価格範囲を1000点でサンプリング
    x_min = max(x1.min(), x2.min())
    x_max = min(x1.max(), x2.max())

    if x_min >= x_max:
        return None, None

    x_common = np.linspace(x_min, x_max, 1000)

    # 線形補間
    y1_interp = np.interp(x_common, x1, y1)
    y2_interp = np.interp(x_common, x2, y2)

    # 交点を探す
    diff = y1_interp - y2_interp
    sign_changes = np.where(np.diff(np.sign(diff)))[0]

    if len(sign_changes) == 0:
        return None, None

    # 最初の交点を返す
    idx = sign_changes[0]
    intersection_x = x_common[idx]
    intersection_y = (y1_interp[idx] + y2_interp[idx]) / 2

    return intersection_x, intersection_y


def create_psm_chart(df_clean: pd.DataFrame, price_columns: dict) -> go.Figure:
    """
    PSM分析のインタラクティブグラフを生成

    Args:
        df_clean: クリーニング済みデータフレーム
        price_columns: 価格列の辞書

    Returns:
        plotly Figure オブジェクト
    """
    fig = go.Figure()

    # 各価格カテゴリの累積分布を計算
    cheap_prices, cheap_cum = calculate_cumulative_distribution(
        df_clean[price_columns['cheap']].values, ascending=False
    )
    bargain_prices, bargain_cum = calculate_cumulative_distribution(
        df_clean[price_columns['bargain']].values, ascending=False
    )
    expensive_prices, expensive_cum = calculate_cumulative_distribution(
        df_clean[price_columns['expensive']].values, ascending=True
    )
    too_expensive_prices, too_expensive_cum = calculate_cumulative_distribution(
        df_clean[price_columns['too_expensive']].values, ascending=True
    )

    # 4本の累積分布曲線を追加
    fig.add_trace(go.Scatter(
        x=cheap_prices,
        y=cheap_cum,
        mode='lines',
        name='安い',
        line=dict(color='blue', width=2),
        hovertemplate='価格: ¥%{x:,.0f}<br>割合: %{y:.1f}%<extra></extra>'
    ))

    fig.add_trace(go.Scatter(
        x=bargain_prices,
        y=bargain_cum,
        mode='lines',
        name='お買い得',
        line=dict(color='green', width=2),
        hovertemplate='価格: ¥%{x:,.0f}<br>割合: %{y:.1f}%<extra></extra>'
    ))

    fig.add_trace(go.Scatter(
        x=expensive_prices,
        y=expensive_cum,
        mode='lines',
        name='高い',
        line=dict(color='orange', width=2),
        hovertemplate='価格: ¥%{x:,.0f}<br>割合: %{y:.1f}%<extra></extra>'
    ))

    fig.add_trace(go.Scatter(
        x=too_expensive_prices,
        y=too_expensive_cum,
        mode='lines',
        name='高すぎる',
        line=dict(color='red', width=2),
        hovertemplate='価格: ¥%{x:,.0f}<br>割合: %{y:.1f}%<extra></extra>'
    ))

    # OPP（最適価格点）の計算と表示
    opp_price, opp_pct = find_intersection(
        expensive_prices, expensive_cum,
        cheap_prices, cheap_cum
    )

    if opp_price is not None:
        fig.add_trace(go.Scatter(
            x=[opp_price],
            y=[opp_pct],
            mode='markers',
            name='OPP（最適価格点）',
            marker=dict(color='purple', size=12, symbol='diamond'),
            hovertemplate='OPP<br>価格: ¥%{x:,.0f}<br>割合: %{y:.1f}%<extra></extra>'
        ))

        # OPPの垂直線
        fig.add_shape(
            type="line",
            x0=opp_price, x1=opp_price,
            y0=0, y1=100,
            line=dict(color="purple", width=1, dash="dash")
        )

    # IDP（妥協価格点）の計算と表示
    idp_price, idp_pct = find_intersection(
        too_expensive_prices, too_expensive_cum,
        bargain_prices, bargain_cum
    )

    if idp_price is not None:
        fig.add_trace(go.Scatter(
            x=[idp_price],
            y=[idp_pct],
            mode='markers',
            name='IDP（妥協価格点）',
            marker=dict(color='brown', size=12, symbol='diamond'),
            hovertemplate='IDP<br>価格: ¥%{x:,.0f}<br>割合: %{y:.1f}%<extra></extra>'
        ))

        # IDPの垂直線
        fig.add_shape(
            type="line",
            x0=idp_price, x1=idp_price,
            y0=0, y1=100,
            line=dict(color="brown", width=1, dash="dash")
        )

    # レイアウト設定
    fig.update_layout(
        title='PSM分析グラフ',
        xaxis_title='価格 (¥)',
        yaxis_title='累積割合 (%)',
        hovermode='closest',
        height=600,
        legend=dict(
            orientation="v",
            yanchor="top",
            y=0.99,
            xanchor="right",
            x=0.99,
            bgcolor="rgba(255, 255, 255, 0.8)"
        )
    )

    return fig


def get_ai_insights(api_key: str, analysis_data: dict, df_clean: pd.DataFrame,
                   price_columns: dict, category_cols: list) -> tuple:
    """
    Claude APIを使用してAI分析を取得

    Args:
        api_key: Anthropic APIキー
        analysis_data: 分析データの辞書
        df_clean: クリーニング済みデータフレーム
        price_columns: 価格列の辞書
        category_cols: カテゴリ列のリスト

    Returns:
        tuple: (AIインサイトのテキスト, エラーメッセージ)
    """
    try:
        # カテゴリ別統計の作成
        category_stats_text = ""
        if category_cols:
            category_stats_text = "\n\n【カテゴリ別統計】\n"
            for cat_col in category_cols:
                category_stats_text += f"\n■ {cat_col}別の平均価格感覚：\n"
                for cat_val in df_clean[cat_col].unique():
                    cat_data = df_clean[df_clean[cat_col] == cat_val]
                    avg_cheap = cat_data[price_columns['cheap']].mean()
                    avg_bargain = cat_data[price_columns['bargain']].mean()
                    avg_expensive = cat_data[price_columns['expensive']].mean()
                    avg_too_expensive = cat_data[price_columns['too_expensive']].mean()
                    category_stats_text += f"  - {cat_val}: 安い=¥{avg_cheap:,.0f}, お買い得=¥{avg_bargain:,.0f}, "
                    category_stats_text += f"高い=¥{avg_expensive:,.0f}, 高すぎる=¥{avg_too_expensive:,.0f}\n"

        # プロンプト構築
        prompt = f"""
以下のPSM（Price Sensitivity Meter）分析結果をもとに、ビジネス向けの包括的なインサイトを提供してください。

【PSM主要指標】
- OPP（最適価格点）: ¥{analysis_data['opp_price']:,.0f}
- IDP（妥協価格点）: ¥{analysis_data['idp_price']:,.0f}
- PMC（最小抵抗価格点）: ¥{analysis_data.get('pmc_price', 'N/A'):,.0f} if isinstance(analysis_data.get('pmc_price'), (int, float)) else 'N/A'
- PME（最大ストレス価格点）: ¥{analysis_data.get('pme_price', 'N/A'):,.0f} if isinstance(analysis_data.get('pme_price'), (int, float)) else 'N/A'

【全体統計】
- 有効回答数: {len(df_clean)}件
- 「安い」価格: 平均¥{df_clean[price_columns['cheap']].mean():,.0f}, 中央値¥{df_clean[price_columns['cheap']].median():,.0f}
- 「お買い得」価格: 平均¥{df_clean[price_columns['bargain']].mean():,.0f}, 中央値¥{df_clean[price_columns['bargain']].median():,.0f}
- 「高い」価格: 平均¥{df_clean[price_columns['expensive']].mean():,.0f}, 中央値¥{df_clean[price_columns['expensive']].median():,.0f}
- 「高すぎる」価格: 平均¥{df_clean[price_columns['too_expensive']].mean():,.0f}, 中央値¥{df_clean[price_columns['too_expensive']].median():,.0f}
{category_stats_text}

以下の形式で、日本語で分析結果を提供してください：

## サマリー
（3行以内で最も重要な発見を簡潔に記述）

## 価格設定の推奨事項
- **推奨価格**: 具体的な金額とその理由
- **推奨理由**:
  - （3-5点の箇条書き）
- **リスクと留意点**:
  - （2-3点の箇条書き）

## 顧客セグメント分析
（カテゴリデータがある場合、セグメント間の違いと各セグメントへの対応策を記述）

## 実装のための具体的なアクションプラン
### 短期（1-3ヶ月）
- （3-5項目）

### 中期（3-6ヶ月）
- （3-5項目）

## 追加調査の提案
- （さらに深掘りすべきポイントを3-5点）

※ビジネスパーソン向けに平易な言葉で、具体的かつ実践的な内容にしてください。
"""

        # Claude API呼び出し
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )

        insights = message.content[0].text
        return insights, None

    except anthropic.AuthenticationError:
        return None, "APIキーが無効です。正しいAPIキーを入力してください。"
    except anthropic.RateLimitError:
        return None, "API制限に達しました。少し待ってから再試行してください。"
    except Exception as e:
        return None, f"AI分析中にエラーが発生しました: {str(e)}"


def create_excel_report(df_clean: pd.DataFrame, price_columns: dict,
                       analysis_data: dict, ai_insights: str = None,
                       category_cols: list = None) -> BytesIO:
    """
    Excelレポートを生成

    Args:
        df_clean: クリーニング済みデータフレーム
        price_columns: 価格列の辞書
        analysis_data: 分析データの辞書
        ai_insights: AIインサイト（オプション）
        category_cols: カテゴリ列のリスト（オプション）

    Returns:
        BytesIO: Excelファイルのバイナリデータ
    """
    wb = Workbook()

    # スタイル定義
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ai_header_fill = PatternFill(start_color="E7F5FF", end_color="E7F5FF", fill_type="solid")

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # シート1: AIインサイト（AI統合版のみ）
    if ai_insights:
        ws_ai = wb.active
        ws_ai.title = "AIインサイト"

        # タイトル
        ws_ai['A1'] = "AI分析レポート"
        ws_ai['A1'].font = Font(bold=True, size=14)
        ws_ai['A1'].fill = ai_header_fill
        ws_ai.merge_cells('A1:D1')

        # AIインサイトを挿入
        ws_ai['A3'] = ai_insights
        ws_ai['A3'].alignment = left_align
        ws_ai.merge_cells('A3:D50')

        # 列幅設定
        ws_ai.column_dimensions['A'].width = 100
        ws_ai.row_dimensions[3].height = 800

    # シート2: 分析サマリー
    ws = wb.create_sheet("分析サマリー")

    row = 1

    # タイトル
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "PSM分析レポート"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    row += 2

    # 分析概要
    ws[f'A{row}'] = "分析日時"
    ws[f'B{row}'] = datetime.now().strftime("%Y年%m月%d日 %H:%M")
    row += 1
    ws[f'A{row}'] = "有効回答数"
    ws[f'B{row}'] = len(df_clean)
    row += 2

    # PSM主要指標
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "PSM主要指標"
    cell.font = Font(bold=True)
    cell.fill = subheader_fill
    row += 1

    metrics = [
        ("OPP（最適価格点）", analysis_data.get('opp_price')),
        ("IDP（妥協価格点）", analysis_data.get('idp_price')),
        ("PMC（最小抵抗価格点）", analysis_data.get('pmc_price')),
        ("PME（最大ストレス価格点）", analysis_data.get('pme_price'))
    ]

    for label, value in metrics:
        ws[f'A{row}'] = label
        if value is not None and isinstance(value, (int, float)):
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '¥#,##0'
            ws[f'B{row}'].fill = highlight_fill
        else:
            ws[f'B{row}'] = "N/A"
        row += 1

    row += 1

    # 全体統計
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "全体統計"
    cell.font = Font(bold=True)
    cell.fill = subheader_fill
    row += 1

    # ヘッダー
    headers = ["項目", "平均", "中央値", "最小値", "最大値", "標準偏差"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    # 統計データ
    price_labels = {
        'cheap': '「安い」価格',
        'bargain': '「お買い得」価格',
        'expensive': '「高い」価格',
        'too_expensive': '「高すぎる」価格'
    }

    for key, label in price_labels.items():
        col = price_columns[key]
        ws[f'A{row}'] = label
        ws[f'B{row}'] = df_clean[col].mean()
        ws[f'C{row}'] = df_clean[col].median()
        ws[f'D{row}'] = df_clean[col].min()
        ws[f'E{row}'] = df_clean[col].max()
        ws[f'F{row}'] = df_clean[col].std()

        for col_idx in range(2, 7):
            ws.cell(row=row, column=col_idx).number_format = '¥#,##0'
            ws.cell(row=row, column=col_idx).border = thin_border
        ws[f'A{row}'].border = thin_border
        row += 1

    # 列幅設定
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15

    # シート3: カテゴリ別分析（オプション）
    if category_cols:
        ws_cat = wb.create_sheet("カテゴリ別分析")
        cat_row = 1

        for cat_col in category_cols:
            # カテゴリタイトル
            ws_cat.merge_cells(f'A{cat_row}:F{cat_row}')
            cell = ws_cat[f'A{cat_row}']
            cell.value = f"{cat_col}別分析"
            cell.font = Font(bold=True, size=12)
            cell.fill = subheader_fill
            cat_row += 1

            # ヘッダー
            headers = [cat_col, "安い(平均)", "お買い得(平均)", "高い(平均)", "高すぎる(平均)", "回答数"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_cat.cell(row=cat_row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            cat_row += 1

            # カテゴリ別データ
            for cat_val in sorted(df_clean[cat_col].unique()):
                cat_data = df_clean[df_clean[cat_col] == cat_val]
                ws_cat[f'A{cat_row}'] = cat_val
                ws_cat[f'B{cat_row}'] = cat_data[price_columns['cheap']].mean()
                ws_cat[f'C{cat_row}'] = cat_data[price_columns['bargain']].mean()
                ws_cat[f'D{cat_row}'] = cat_data[price_columns['expensive']].mean()
                ws_cat[f'E{cat_row}'] = cat_data[price_columns['too_expensive']].mean()
                ws_cat[f'F{cat_row}'] = len(cat_data)

                for col_idx in range(2, 6):
                    ws_cat.cell(row=cat_row, column=col_idx).number_format = '¥#,##0'
                    ws_cat.cell(row=cat_row, column=col_idx).border = thin_border
                ws_cat[f'A{cat_row}'].border = thin_border
                ws_cat[f'F{cat_row}'].border = thin_border
                cat_row += 1

            cat_row += 2  # カテゴリ間の空白

        # 列幅設定
        ws_cat.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_cat.column_dimensions[col].width = 18

    # シート4: 生データ
    ws_data = wb.create_sheet("生データ")

    # ヘッダー
    for col_idx, col_name in enumerate(df_clean.columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # データ
    for row_idx, row_data in enumerate(df_clean.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_data.cell(row=row_idx, column=col_idx).value = value

    # BytesIOに保存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


# ========== メインUI ==========

# パスワード認証チェック
if not check_password():
    st.stop()

# サイドバー
with st.sidebar:
    st.title("📊 PSM分析ツール（AI版）")

    st.markdown("---")

    # API Key設定状況の表示
    if st.session_state.api_key:
        st.success("✅ APIキーが設定されています")
    else:
        st.error("❌ APIキーが設定されていません\n.envファイルにAPIキーを設定してください")

    st.markdown("---")

    # ファイルアップロード
    st.subheader("📁 データアップロード")
    uploaded_file = st.file_uploader(
        "CSVファイルを選択",
        type=['csv'],
        help="PSMアンケートデータのCSVファイルをアップロードしてください"
    )

    st.markdown("---")

    # 使い方
    with st.expander("📖 使い方", expanded=False):
        st.markdown("""
        ### 基本的な使い方

        1. **APIキーを入力**
           - Anthropic Claude APIキーを取得
           - 上記の入力欄に貼り付け

        2. **CSVファイルをアップロード**
           - PSMアンケートデータを用意
           - ファイルを選択してアップロード

        3. **列を選択**
           - 4つの必須価格列を選択
           - オプションでカテゴリ列を選択

        4. **分析を実行**
           - ボタンをクリックして分析開始
           - AI分析が完了するまで待機

        5. **結果を確認・ダウンロード**
           - グラフとインサイトを確認
           - Excelレポートをダウンロード
        """)

    # リンク
    with st.expander("🔗 リンク", expanded=False):
        st.markdown("""
        - [Anthropic API](https://console.anthropic.com/)
        - [PSM分析について](https://www.macromill.com/service/data-analysis/price-sensitivity-meter/)
        - [サンプルデータ](./sample_data.csv)
        """)

    st.markdown("---")

    # ログアウトボタン
    if st.button("🚪 ログアウト", use_container_width=True):
        st.session_state.authenticated = False
        st.rerun()

# メインエリア
if uploaded_file is not None:
    try:
        # CSVファイルの読み込み
        try:
            df = pd.read_csv(uploaded_file, encoding='utf-8')
        except:
            df = pd.read_csv(uploaded_file, encoding='shift-jis')

        st.success(f"✅ ファイルを読み込みました（{len(df)}行 × {len(df.columns)}列）")

        # データプレビュー
        with st.expander("📋 データプレビュー", expanded=False):
            st.dataframe(df.head(10), use_container_width=True)

        st.markdown("---")

        # 列選択セクション
        st.subheader("🎯 列の選択")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### 必須列（価格）")

            cheap_col = st.selectbox(
                "「安い」と感じる価格",
                options=df.columns.tolist(),
                help="この価格だと「安い」と感じる価格の列を選択"
            )

            bargain_col = st.selectbox(
                "「お買い得」と感じる価格",
                options=df.columns.tolist(),
                help="この価格だと「お買い得」と感じる価格の列を選択"
            )

            expensive_col = st.selectbox(
                "「高い」と感じる価格",
                options=df.columns.tolist(),
                help="この価格だと「高い」と感じる価格の列を選択"
            )

            too_expensive_col = st.selectbox(
                "「高すぎる」と感じる価格",
                options=df.columns.tolist(),
                help="この価格だと「高すぎる」と感じる価格の列を選択"
            )

        with col2:
            st.markdown("#### オプション列（カテゴリ）")

            category_cols = st.multiselect(
                "カテゴリ列（複数選択可）",
                options=[col for col in df.columns.tolist()
                        if col not in [cheap_col, bargain_col, expensive_col, too_expensive_col]],
                help="性別、年齢、利用頻度など、セグメント分析したい列を選択"
            )

            st.info("カテゴリ列を選択すると、セグメント別の分析が可能になります")

        # 選択した列のプレビュー
        if st.checkbox("選択列のプレビューを表示", value=False):
            preview_cols = [cheap_col, bargain_col, expensive_col, too_expensive_col] + category_cols
            st.dataframe(df[preview_cols].head(), use_container_width=True)

        st.markdown("---")

        # 分析実行ボタン
        analyze_button = st.button("🚀 分析を実行", type="primary", use_container_width=True)

        if analyze_button:
            # バリデーション
            if not st.session_state.api_key:
                st.error("❌ APIキーを入力してください")
                st.stop()

            # 価格列の辞書
            price_columns = {
                'cheap': cheap_col,
                'bargain': bargain_col,
                'expensive': expensive_col,
                'too_expensive': too_expensive_col
            }

            # データクリーニング
            with st.spinner("データをクリーニング中..."):
                df_clean = df.copy()

                # 価格列を数値に変換
                for col in price_columns.values():
                    df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')

                # 欠損値を削除
                df_clean = df_clean.dropna(subset=list(price_columns.values()))

                # データ検証
                if len(df_clean) < 10:
                    st.error(f"❌ 分析には最低10件のデータが必要です（有効データ: {len(df_clean)}件）")
                    st.stop()

                st.success(f"✅ データクリーニング完了（有効データ: {len(df_clean)}件）")

            # PSM分析実行
            with st.spinner("PSM分析を実行中..."):
                # 累積分布の計算
                cheap_prices, cheap_cum = calculate_cumulative_distribution(
                    df_clean[cheap_col].values, ascending=False
                )
                bargain_prices, bargain_cum = calculate_cumulative_distribution(
                    df_clean[bargain_col].values, ascending=False
                )
                expensive_prices, expensive_cum = calculate_cumulative_distribution(
                    df_clean[expensive_col].values, ascending=True
                )
                too_expensive_prices, too_expensive_cum = calculate_cumulative_distribution(
                    df_clean[too_expensive_col].values, ascending=True
                )

                # 主要指標の計算
                opp_price, opp_pct = find_intersection(
                    expensive_prices, expensive_cum,
                    cheap_prices, cheap_cum
                )

                idp_price, idp_pct = find_intersection(
                    too_expensive_prices, too_expensive_cum,
                    bargain_prices, bargain_cum
                )

                pmc_price, pmc_pct = find_intersection(
                    expensive_prices, expensive_cum,
                    bargain_prices, bargain_cum
                )

                pme_price, pme_pct = find_intersection(
                    too_expensive_prices, too_expensive_cum,
                    cheap_prices, cheap_cum
                )

                analysis_data = {
                    'opp_price': opp_price,
                    'idp_price': idp_price,
                    'pmc_price': pmc_price,
                    'pme_price': pme_price
                }

                st.success("✅ PSM分析完了")

            # AI分析実行
            with st.spinner("🤖 AIがデータを分析中...（30秒程度かかります）"):
                ai_insights, error = get_ai_insights(
                    st.session_state.api_key,
                    analysis_data,
                    df_clean,
                    price_columns,
                    category_cols
                )

                if error:
                    st.error(f"❌ {error}")
                    ai_insights = None
                else:
                    st.success("✅ AI分析完了")

            # セッション状態に保存
            st.session_state.analysis_complete = True
            st.session_state.df_clean = df_clean
            st.session_state.price_columns = price_columns
            st.session_state.analysis_data = analysis_data
            st.session_state.ai_insights = ai_insights
            st.session_state.category_cols = category_cols

            st.markdown("---")

            # 結果表示
            st.header("📊 分析結果")

            # AIインサイト（ハイライトボックス）
            if ai_insights:
                # f-string内でバックスラッシュを使用できないため、事前に変換
                ai_insights_html = ai_insights.replace('\n', '<br>')
                st.markdown(f"""
                <div style="
                    background-color: #F8FCFF;
                    padding: 20px;
                    border-radius: 10px;
                    border-left: 5px solid #0066CC;
                    margin-bottom: 20px;
                    color: #2C3E50;
                ">
                <h3 style="color: #0066CC; margin-top: 0;">🤖 AIインサイト</h3>
                <div style="color: #2C3E50; line-height: 1.6;">
                {ai_insights_html}
                </div>
                </div>
                """, unsafe_allow_html=True)

            # 主要指標（メトリクスカード）
            st.subheader("📈 PSM主要指標")

            col1, col2, col3, col4 = st.columns(4)

            with col1:
                if opp_price is not None:
                    st.metric(
                        label="OPP（最適価格点）",
                        value=f"¥{int(opp_price):,}",
                        help="「高い」と「安い」の交点"
                    )
                else:
                    st.metric(label="OPP", value="N/A")

            with col2:
                if idp_price is not None:
                    st.metric(
                        label="IDP（妥協価格点）",
                        value=f"¥{int(idp_price):,}",
                        help="「高すぎる」と「お買い得」の交点"
                    )
                else:
                    st.metric(label="IDP", value="N/A")

            with col3:
                if pmc_price is not None:
                    st.metric(
                        label="PMC（最小抵抗価格点）",
                        value=f"¥{int(pmc_price):,}",
                        help="「高い」と「お買い得」の交点"
                    )
                else:
                    st.metric(label="PMC", value="N/A")

            with col4:
                if pme_price is not None:
                    st.metric(
                        label="PME（最大ストレス価格点）",
                        value=f"¥{int(pme_price):,}",
                        help="「高すぎる」と「安い」の交点"
                    )
                else:
                    st.metric(label="PME", value="N/A")

            st.markdown("---")

            # グラフ
            st.subheader("📉 PSM分析グラフ")
            fig = create_psm_chart(df_clean, price_columns)
            st.plotly_chart(fig, use_container_width=True)

            st.markdown("---")

            # 統計テーブル
            st.subheader("📊 全体統計")

            stats_data = []
            for key, label in [('cheap', '「安い」価格'),
                              ('bargain', '「お買い得」価格'),
                              ('expensive', '「高い」価格'),
                              ('too_expensive', '「高すぎる」価格')]:
                col = price_columns[key]
                stats_data.append({
                    '項目': label,
                    '平均': f"¥{int(df_clean[col].mean()):,}",
                    '中央値': f"¥{int(df_clean[col].median()):,}",
                    '最小値': f"¥{int(df_clean[col].min()):,}",
                    '最大値': f"¥{int(df_clean[col].max()):,}",
                    '標準偏差': f"¥{int(df_clean[col].std()):,}"
                })

            st.dataframe(pd.DataFrame(stats_data), use_container_width=True, hide_index=True)

            # カテゴリ別分析
            if category_cols:
                st.markdown("---")
                with st.expander("📊 カテゴリ別分析", expanded=True):
                    for cat_col in category_cols:
                        st.markdown(f"#### {cat_col}別の分析")

                        cat_stats = []
                        for cat_val in sorted(df_clean[cat_col].unique()):
                            cat_data = df_clean[df_clean[cat_col] == cat_val]
                            cat_stats.append({
                                cat_col: cat_val,
                                '安い(平均)': f"¥{int(cat_data[cheap_col].mean()):,}",
                                'お買い得(平均)': f"¥{int(cat_data[bargain_col].mean()):,}",
                                '高い(平均)': f"¥{int(cat_data[expensive_col].mean()):,}",
                                '高すぎる(平均)': f"¥{int(cat_data[too_expensive_col].mean()):,}",
                                '回答数': len(cat_data)
                            })

                        st.dataframe(pd.DataFrame(cat_stats), use_container_width=True, hide_index=True)
                        st.markdown("")

            st.markdown("---")

            # Excelダウンロード
            st.subheader("📥 レポートダウンロード")

            with st.spinner("Excelレポートを生成中..."):
                excel_data = create_excel_report(
                    df_clean,
                    price_columns,
                    analysis_data,
                    ai_insights,
                    category_cols
                )

            st.download_button(
                label="📥 Excelレポートをダウンロード",
                data=excel_data,
                file_name=f"PSM分析レポート_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

    except Exception as e:
        st.error(f"❌ エラーが発生しました: {str(e)}")
        st.exception(e)

else:
    # 初期画面
    st.title("📊 PSM分析自動レポートツール（AI版）")

    st.markdown("""
    ### Claude AIを活用した高度なPSM分析

    このツールは、Van WestendorpのPrice Sensitivity Meter（PSM）分析を自動化し、
    Anthropic Claude APIによる深いインサイトを提供します。

    #### 主な機能

    - 📊 **自動PSM分析**: 4つの価格ポイント（OPP、IDP、PMC、PME）を自動計算
    - 🤖 **AI分析**: Claude APIによる包括的なビジネスインサイト生成
    - 📈 **インタラクティブグラフ**: Plotlyによる美しい可視化
    - 📑 **Excelレポート**: AI分析を含む詳細なレポートを自動生成
    - 🎯 **カテゴリ別分析**: セグメント別の詳細な価格感覚を分析

    #### 使い方

    1. **左サイドバーでAPIキーを入力**
    2. **CSVファイルをアップロード**
    3. **価格列とカテゴリ列を選択**
    4. **「分析を実行」ボタンをクリック**
    5. **結果を確認してExcelレポートをダウンロード**

    #### 必要なデータ形式

    CSVファイルに以下の4つの価格列が必要です：
    - 「安い」と感じる価格
    - 「お買い得」と感じる価格
    - 「高い」と感じる価格
    - 「高すぎる」と感じる価格

    ---

    **👈 左サイドバーからファイルをアップロードして開始してください**
    """)

    # サンプルデータの表示
    if os.path.exists('sample_data.csv'):
        with st.expander("📋 サンプルデータを見る", expanded=False):
            try:
                sample_df = pd.read_csv('sample_data.csv')
                st.dataframe(sample_df.head(), use_container_width=True)
                st.info("このようなフォーマットのCSVファイルをアップロードしてください")
            except:
                pass
